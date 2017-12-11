#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import time,datetime
import csv
import xlrd,xlsxwriter
import openpyxl

from model import AmzTransactionInfo,BaseMethod

_MONTH_2_NO_Eg = {
	"Jan":"01",
	"Feb":"02",
	"Mar":"03",
	"Apr":"04",
	"May":"05",
	"Jun":"06",
	"Jul":"07",
	"Aug":"08",
	"Sep":"09",
	"Oct":"10",
	"Nov":"11",
	"Dec":"12"
}

_MONTH_2_NO_Fr = {
	"janv":"01",
	"févr":"02",
	"mars":"03",
	"avr":"04",
	"mai":"05",
	"juin":"06",
	"juil":"07",
	"août":"08",
	"sept":"09",
	"oct":"10",
	"nov.":"11",
	"déc.":"12"
}
_MONTH_2_NO_It = {
	"gen":"01",
	"feb":"02",
	"mar":"03",
	"apr":"04",
	"mag":"05",
	"giu":"06",
	"lug":"07",
	"ago":"08",
	"set":"09",
	"ott":"10",
	"nov":"11",
	"dic":"12"
}
_STATE_2_AMZ = {
	"ye":"amazon.com",
	"us":"amazon.com",
	"ca":"amazon.ca",
	"uk":"amazon.co.uk",
	"de":"amazon.de",
	"fr":"amazon.fr",
	"it":"amazon.it",
	"es":"amazon.es"
}
_TYPE = {
	"A-to-z Guarantee Claim":["A-to-z Guarantee Claim",],
	"Adjustment":["Adjustment","Anpassung","Ajustement","Ajuste"],
	"Debt":["Debt",],
	"FBA Customer Return Fee":["FBA Customer Return Fee",
							   "Erstattung durch Rückbuchung"
							  ],
	"FBA Inventory Fee":["FBA Inventory Fee",
						 "Versand durch Amazon Lagergebühr",
						 "Frais de stock Expédié par Amazon",
						 "Costo di stoccaggio Logistica di Amazon",
						 "Tarifas de inventario de Logística de Amazon"
						],
	"Order":["Order","Bestellung","Commande","Ordine","Pedido"],
	"Refund":["Refund","Erstattung","Remboursement","Rimborso","Reembolso"],
	"Service Fee":["Service Fee","Servicegebühr",],
	"Transfer":["Transfer","Übertrag","Transfert","Trasferimento","Transferir"],
	"Lightning Deal Fee":["Lightning Deal Fee",
						  "Blitzangebotsgebühr",
						  "Tarif de la Vente Flash",
						  "Tarifa de Oferta flash"
						 ]

}

_North_America = ["ye","us","ca"]

_EU = ["uk","de","fr","it","es"]

_STATE_LIST = ["amazon.com","amazon.ca","amazon.co.uk","amazon.de","amazon.fr","amazon.it","amazon.es"]
def transaction_info_into_database():

	def _formate_datetime(state,date_time):

		def _formate_datetime_fr(date_time):
			blanks_list_of_datetime = []
			i = 0
			for i in range(0,len(date_time)+1):
				if date_time[i:i+1] == " " :
					blanks_list_of_datetime.append(i)
			return blanks_list_of_datetime

		if state in ("ye","us"):
			if date_time.find(',') < 6:
				date_time = date_time[:4]+"0"+date_time[4:]
			if date_time.find(':') < 15:
				date_time = date_time[:13]+"0"+date_time[13:]
				print(date_time[13:15])
			if date_time.find("PM"):
				if int(date_time[13:15]) >= 12:
					date_time_H = int(date_time[13:15])
				else:
					date_time_H = int(date_time[13:15])+12
				# don't use replace()
				date_time = date_time[:13]+str(int(date_time[13:15])+12)+date_time[15:]
				print(date_time[13:15])
			if date_time.find("AM"):
				if int(date_time[13:15]) >= 12:
					date_time_H = int(date_time[13:15])-12
				else:
					date_time_H = int(date_time[13:15])
				date_time = date_time[:13]+str(date_time_H)+date_time[15:]
			datetime = date_time[8:12]+'-'+_MONTH_2_NO_Eg[date_time[:3]]+'-'+date_time[4:6]+date_time[12:21]

		if state in ("ca"):
			if date_time.find(':') < 13:
				date_time = date_time[:11]+"0"+date_time[11:]
			if date_time.find("PM"):
				if int(date_time[11:13]) >= 12:
					date_time_H = int(date_time[11:13])
				else:
					date_time_H = int(date_time[11:13])+12
				# don't use replace()
				datetime = date_time[:11]+str(date_time_H)+date_time[13:]
			if date_time.find("AM"):
				if int(date_time[11:13]) >= 12:
					date_time_H = int(date_time[11:13])-12
				else:
					date_time_H = int(date_time[11:13])
				datetime = date_time[:11]+str(date_time_H)+date_time[13:]

		if state in ("uk"):
			if date_time.find(":") < 14:
				date_time = "0"+date_time
			datetime = date_time[7:11]+'-'+_MONTH_2_NO_Eg[date_time[3:6]]+'-'+date_time[:2]+date_time[11:date_time.find("GMT")-1]


		if state in ("de","es"):
			datetime = date_time[6:10]+'-'+date_time[3:5]+'-'+date_time[:2]+date_time[10:date_time.find('GMT')-1]

		if state in ("it"):
			datetime = date_time[7:11]+'-'+_MONTH_2_NO_It[date_time[3:6]]+'-'+date_time[:2]+date_time[11:date_time.find('GMT')-1].replace('.',':')

		if state in ("fr"):
			if date_time.find(" ") <2:
				date_time="0"+date_time
				blanks_list_of_datetime = _formate_datetime_fr(date_time)
			else:
				blanks_list_of_datetime = _formate_datetime_fr(date_time)
			datetime = (date_time[blanks_list_of_datetime[1]+1:blanks_list_of_datetime[1]+5]+'-'+
						_MONTH_2_NO_Fr[date_time[blanks_list_of_datetime[0]:blanks_list_of_datetime[1]].replace(".","").strip()]+'-'+
						date_time[:blanks_list_of_datetime[0]]+
						date_time[blanks_list_of_datetime[2]:date_time.find('UTC')-1].replace('.',':'))

		return datetime
		
	def _formate_money(state, money):
		if state in ("ye","us","ca","uk") :
			money = money.replace(',','')
		if state in ("de","fr","it","es"):
			money = money.replace(u'\xa0', u'').replace('.','').replace(',','.')
		return money

	def _formate_type(type):
		if type in _TYPE["A-to-z Guarantee Claim"] :
			type = "A-to-z Guarantee Claim"
		if type in _TYPE["Adjustment"] :
			type = "Adjustment"
		if type in _TYPE["Debt"] :
			type = "Debt"
		if type in _TYPE["FBA Customer Return Fee"] :
			type = "FBA Customer Return Fee"
		if type in _TYPE["FBA Inventory Fee"] :
			type = "FBA Inventory Fee"
		if type in _TYPE["Order"] :
			type = "Order"
		if type in _TYPE["Refund"] :
			type = "Refund"
		if type in _TYPE["Service Fee"] :
			type = "Service Fee"
		if type in _TYPE["Transfer"] :
			type = "Transfer"
		if type in _TYPE["Lightning Deal Fee"] :
			type = "Lightning Deal Fee"

		return type	

	def _process_file(state):
		#read the every state csv and write into database
		with open(file='C:\\Users\\ACEEC\\Desktop\\amz_payment\\2017AugMonthlyTransaction_%s.csv' % state,
				  mode='r',
				  encoding='utf-8'
				  ) as csv_file:

			reader = csv.reader(csv_file)
			print(type(reader))
			#type of reader is _csv.reader
			NO = 1 
			rows = [row for row in reader]

			# first of row North Amecia different with EU
			if state in _North_America:
				state_rows = rows[8:]
			if state in _EU:
				state_rows = rows[7:]

			for row in state_rows:
				amz_trans_info = AmzTransactionInfo()
				amz_trans_info.date_time = _formate_datetime(state,row[0])
				amz_trans_info.settlement_id = row[1]
				# unify as English
				amz_trans_info.type = _formate_type(row[2])
				amz_trans_info.order_id = row[3]
				amz_trans_info.sku = row[4]
				amz_trans_info.description = row[5]
				#some of type's quantity is null but databsa doesn't allow to be null
				if row[6]:
					amz_trans_info.quantity = row[6]
				else:
					amz_trans_info.quantity = -1
				# marketplace must be not null and must modify and unify as lower_case
				if row[7]:
					amz_trans_info.marketplace = row[7]
				else:
					amz_trans_info.marketplace = _STATE_2_AMZ[state]
				amz_trans_info.fulfillment_channel = row[8]
				amz_trans_info.order_city = row[9]
				amz_trans_info.order_state = row[10]
				amz_trans_info.order_postal = row[11]
				amz_trans_info.product_sales = _formate_money(state,row[12])
				amz_trans_info.shipping_credits = _formate_money(state,row[13])
				amz_trans_info.gift_wrap_credits = _formate_money(state,row[14])
				amz_trans_info.promotional_rebates = _formate_money(state,row[15])
				#de not have sales_tax_collected
				if  state in _EU:
					amz_trans_info.sales_tax_collected = 0
					amz_trans_info.selling_fees = _formate_money(state,row[16])
					amz_trans_info.fba_fees = _formate_money(state,row[17])
					amz_trans_info.other_transaction_fees = _formate_money(state,row[18])
					amz_trans_info.other = _formate_money(state,row[19])
					amz_trans_info.total = _formate_money(state,row[20])
				if  state in _North_America:
					amz_trans_info.sales_tax_collected = row[16]
					amz_trans_info.selling_fees = _formate_money(state,row[17])
					amz_trans_info.fba_fees = _formate_money(state,row[18])
					amz_trans_info.other_transaction_fees = _formate_money(state,row[19])
					amz_trans_info.other = _formate_money(state,row[20])
					amz_trans_info.total = _formate_money(state,row[21])

				amz_trans_info.add(False)
				#just for check out
				print("the %s record" % NO)
				NO += 1
			BaseMethod.commit()
			print('--------------------------------byebye :transaction details of %s write to database is done !------------------------' % state)
	 	
	# # just for unit test
	#_process_file("es")
	
	states = ["ye","us","ca","uk","de","fr","it","es"]
	try:
		for state in states:
			try:
				_process_file(state)
			except Exception as e:
				print("------------------------%s transaction writing into database failure------------------------" % state )
				raise
				continue
			else:
				time.sleep(5)
	except Exception as e:
		raise
		os._exit(0)




# process xlsx by openpysxl
def monthly_sku_qty_profit_statistics_new():

	def _list_transaction_data(amz_trans_info):
		trans_list = [
					  amz_trans_info.sku,
					  amz_trans_info.quantity,
					  amz_trans_info.other,
					  amz_trans_info.total,
					  amz_trans_info.fulfillment_channel,
					  amz_trans_info.marketplace
					 ]
		return trans_list

	def _list_other_expenses(otherExpensesInfo):
		expenses_List = [
						 otherExpensesInfo.type,
						 "\t",
						 "\t",
						 otherExpensesInfo.expenses,
						 otherExpensesInfo.marketplace,
						]
		return expenses_List

	_TITLE_LIST = ["sku","quantity","total","profit","fulfillment_channel","station"]
	_EURO_ZONE = ['amazon.de','amazon.fr','amazon.it','amazon.es'] 

	#creat List to collect diffrent states' transacion record
	COMList = []
	CAList  = []
	DEList  = []
	FRList  = []
	ITList  = []
	ESList  = []
	EUList  = []
	UKList = []
	# Classify and statistics of sku,quantity,profits
	TransInfo = AmzTransactionInfo.accounting_sku_qty_profit_statistics()
	for transInfo in TransInfo:
		amz_trans_info = AmzTransactionInfo()
		amz_trans_info.fulfillment_channel = transInfo.fulfillment_channel
		amz_trans_info.sku = transInfo.sku
		amz_trans_info.quantity = transInfo.quantity
		amz_trans_info.other = transInfo.total
		amz_trans_info.total = transInfo.profit
		amz_trans_info.marketplace = transInfo.marketplace

		if amz_trans_info.marketplace == 'amazon.com':
			COMList.append(_list_transaction_data(amz_trans_info))
		if amz_trans_info.marketplace == 'amazon.ca':
			CAList.append(_list_transaction_data(amz_trans_info))
		if amz_trans_info.marketplace in ('amazon.co.uk'):
			UKList.append(_list_transaction_data(amz_trans_info))
		if amz_trans_info.marketplace == 'amazon.de':
			DEList.append(amz_trans_info)
		if amz_trans_info.marketplace == 'amazon.fr':
			FRList.append(amz_trans_info)
		if amz_trans_info.marketplace == 'amazon.it':
			ITList.append(amz_trans_info)
		if amz_trans_info.marketplace == 'amazon.es':
			ESList.append(amz_trans_info)


	#creat List to collect diffrent states' expenses record
	COMEXPENList = []
	CAEXPENList = []
	UKEXPENList = []
	DEEXPENList = []
	FREXPENList = []
	ITEXPENList = []
	ESEXPENList = []
	#classify and statistic of other expenses
	OtherExpensesInfo = AmzTransactionInfo.accounting_other_expensess_statistics()
	for otherExpensesInfo in OtherExpensesInfo:

		if otherExpensesInfo.marketplace.lower() == "amazon.com":
			COMEXPENList.append(_list_other_expenses(otherExpensesInfo))
		if otherExpensesInfo.marketplace.lower() == "amazon.ca":
			CAEXPENList.append(_list_other_expenses(otherExpensesInfo))
		if otherExpensesInfo.marketplace.lower() == "amazon.co.uk":
			UKEXPENList.append(_list_other_expenses(otherExpensesInfo))
		if otherExpensesInfo.marketplace.lower() == "amazon.de":
			DEEXPENList.append(_list_other_expenses(otherExpensesInfo))
		if otherExpensesInfo.marketplace.lower() == "amazon.fr":
			FREXPENList.append(_list_other_expenses(otherExpensesInfo))
		if otherExpensesInfo.marketplace.lower() == "amazon.it":
			ITEXPENList.append(_list_other_expenses(otherExpensesInfo))
		if otherExpensesInfo.marketplace.lower() == "amazon.es":
			ESEXPENList.append(_list_other_expenses(otherExpensesInfo))


	Workbook = openpyxl.Workbook()
	if len(COMList)!=0:
			com_sheet = Workbook.create_sheet('com')
			com_sheet.append(_TITLE_LIST)
			COMrows = len(COMList)
			COMcols = len(COMList[0])
			for comRow in range(0,COMrows):
				for comCol in range(0,COMcols):
					if COMList[0][5] == 'amazon.com':
						com_sheet.cell(row=comRow+2,column=comCol+1,value=COMList[comRow][comCol])
					else:
						continue
			for com in COMEXPENList:
				com_sheet.append(com)
	if len(CAList)!=0:
			ca_sheet = Workbook.create_sheet('CA')
			ca_sheet.append(_TITLE_LIST)
			CArows = len(CAList)
			CAcols = len(CAList[0])
			for caRow in range(0,CArows):
				for caCol in range(0,CAcols):
					if CAList[0][5] == 'amazon.ca':
						ca_sheet.cell(row=caRow+2,column=caCol+1,value=CAList[caRow][caCol])
					else:
						continue
			for ca in CAEXPENList:
				ca_sheet.append(ca)

	# add the other station's sku which is not in germans into DEList
	for frList in FRList:
		if frList.sku  not in [deList.sku for deList in DEList]:
			DEList.append(frList)
	for itList in ITList:
		if itList.sku not in [deList.sku for deList in DEList]:
			DEList.append(itList)
	for esList in ESList:
		if esList.sku not in [deList.sku for deList in DEList]:
			DEList.append(esList)

	# add up quantity,total,profit of commen sku
	for deList in DEList:
		for frList in FRList:
			if  deList.sku == frList.sku and deList.marketplace != frList.marketplace:
				deList.quantity+=frList.quantity
				deList.other+=frList.other
				deList.total+=frList.total
		for itList in ITList:
			if  deList.sku == itList.sku and deList.marketplace != itList.marketplace:
				deList.quantity+=itList.quantity
				deList.other+=itList.other
				deList.total+=itList.total	
		for esList in ESList:
			if  deList.sku == esList.sku and deList.marketplace != esList.marketplace:
				deList.quantity+=esList.quantity
				deList.other+=esList.other
				deList.total+=esList.total
		eulist = [deList.sku,
				  deList.quantity,
				  deList.other,
				  deList.total,
				  deList.fulfillment_channel,
				  deList.marketplace
				 ]
		EUList.append(eulist)

	
		

	if len(EUList) != 0:
		eu_sheet = Workbook.create_sheet('EU')
		eu_sheet.append(_TITLE_LIST)
		EUrows = len(EUList)
		EUcols = len(EUList[0])
		for euRow in range(0,EUrows):
			for euCol in range(0,EUcols):
				if EUList[0][5] in _EURO_ZONE:
					eu_sheet.cell(row=euRow+2,column=euCol+1,value=EUList[euRow][euCol])
				else:
					continue
		for de in DEEXPENList:
			eu_sheet.append(de)
		for fr in FREXPENList:
			eu_sheet.append(fr)
		for it in ITEXPENList:
			eu_sheet.append(it)
		for es in ESEXPENList:
			eu_sheet.append(es)

	if len(UKList) != 0:
		uk_sheet = Workbook.create_sheet('UK')
		uk_sheet.append(_TITLE_LIST)
		# print(len(UKList))
		# print(len(UKList[0]))
		Ukrows = len(UKList)
		Ukcols = len(UKList[0])
		for ukRow in range(0,Ukrows):
			for ukCol in range(0,Ukcols):
				if UKList[0][5] == 'amazon.co.uk':
					uk_sheet.cell(row=ukRow+2,column=ukCol+1,value=UKList[ukRow][ukCol])
				else:
					continue
		for uk in UKEXPENList:
			uk_sheet.append(uk)

	Workbook.save('./%s-%s_new_test.xlsx' % ((datetime.datetime.now().year,datetime.datetime.now().month-1)))
	print('byebye : Classify and statistics id done !')
	


if __name__ == '__main__':

	transaction_info_into_database()
	#monthly_sku_qty_profit_statistics_new()